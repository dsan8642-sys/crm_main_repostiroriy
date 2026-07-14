"""Shared export helpers (module 5.9/5.10). Produce bytes for .xlsx and CSV.

CSV uses UTF-8 BOM so Cyrillic opens correctly in Excel. XLSX uses a professional
font with a bold, frozen header row. These are data exports (no formulas)."""
import csv
import io

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

FONT = "Arial"
HEADER_FILL = PatternFill("solid", start_color="1F4E78")
HEADER_FONT = Font(name=FONT, bold=True, color="FFFFFF")
BODY_FONT = Font(name=FONT)


def rows_to_csv(headers, rows) -> bytes:
    buf = io.StringIO()
    writer = csv.writer(buf, delimiter=";")  # PL locale friendly
    writer.writerow(headers)
    for r in rows:
        writer.writerow(["" if v is None else v for v in r])
    return ("\ufeff" + buf.getvalue()).encode("utf-8")


def _fill_sheet(ws, headers, rows):
    ws.append(headers)
    for col, _ in enumerate(headers, start=1):
        c = ws.cell(row=1, column=col)
        c.font = HEADER_FONT
        c.fill = HEADER_FILL
        c.alignment = Alignment(vertical="center")
    for r in rows:
        ws.append(["" if v is None else v for v in r])
        for col in range(1, len(headers) + 1):
            ws.cell(row=ws.max_row, column=col).font = BODY_FONT
    # width heuristic
    for col, header in enumerate(headers, start=1):
        width = max([len(str(header))] + [len(str(r[col - 1])) for r in rows] + [8]) + 2
        ws.column_dimensions[get_column_letter(col)].width = min(width, 50)
    ws.freeze_panes = "A2"


def rows_to_xlsx(sheet_title, headers, rows) -> bytes:
    wb = Workbook()
    ws = wb.active
    ws.title = sheet_title[:31]
    _fill_sheet(ws, headers, rows)
    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def sheets_to_xlsx(sheets) -> bytes:
    """sheets: list of (title, headers, rows) -> multi-sheet workbook bytes."""
    wb = Workbook()
    wb.remove(wb.active)
    for title, headers, rows in sheets:
        ws = wb.create_sheet(title[:31])
        _fill_sheet(ws, headers, rows)
    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()
