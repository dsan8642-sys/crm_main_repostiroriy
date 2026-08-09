"""Module 5.10: import clients from .xlsx/.csv — parse, map, preview, commit, rollback."""
import csv
import io
from zipfile import BadZipFile
from dataclasses import dataclass, field
from datetime import date

from django.core.exceptions import ValidationError
from django.core.validators import validate_email
from django.db import transaction
from django.utils import timezone

from accounts.models import ParentAccount, User
from audit.models import audit
from catalog.models import Group, SubscriptionType
from students.models import Student
from subscriptions.models import allow_subscription_history_delete
from subscriptions.services import create_subscription

from .contracts import prepare_rows
from .models import ImportBatch, ImportBatchStatus, ImportKind

CANONICAL_FIELDS = ("first_name", "last_name", "name", "phone", "email", "group", "subscription")
MAX_IMPORT_BYTES = 5 * 1024 * 1024
MAX_IMPORT_ROWS = 5000
MAX_IMPORT_COLUMNS = 50

NEW, DUPLICATE, ERROR = "new", "duplicate", "error"


@dataclass
class PreviewRow:
    index: int
    data: dict
    status: str = NEW
    errors: list = field(default_factory=list)
    warnings: list = field(default_factory=list)
    resolved: dict = field(default_factory=dict)


def _decode_csv(raw: bytes) -> str:
    for enc in ("utf-8-sig", "cp1250", "iso-8859-2", "utf-8"):  # PL-friendly fallbacks
        try:
            return raw.decode(enc)
        except UnicodeDecodeError:
            continue
    return raw.decode("utf-8", errors="replace")


def parse_source(raw: bytes, filename: str):
    """Return (headers, list_of_row_dicts) from .xlsx or .csv bytes."""
    if len(raw) > MAX_IMPORT_BYTES:
        raise ValidationError("Файл импорта превышает лимит 5 МБ")
    lower_name = filename.lower()
    if not lower_name.endswith((".xlsx", ".xlsm", ".csv")):
        raise ValidationError("Поддерживаются только CSV, XLSX и XLSM")
    if lower_name.endswith((".xlsx", ".xlsm")):
        from openpyxl import load_workbook
        from openpyxl.utils.exceptions import InvalidFileException
        try:
            wb = load_workbook(io.BytesIO(raw), read_only=True, data_only=False)
        except (BadZipFile, InvalidFileException, OSError, ValueError) as exc:
            raise ValidationError("Повреждённый или неподдерживаемый XLSX-файл") from exc
        ws = wb.active
        row_iterator = ws.iter_rows()
        try:
            header_cells = next(row_iterator)
        except StopIteration:
            return [], []
        if any(cell.data_type == "f" for cell in header_cells):
            raise ValidationError("Формулы в import-файле запрещены")
        header_row = [cell.value for cell in header_cells]
        if len(header_row) > MAX_IMPORT_COLUMNS:
            raise ValidationError("Файл импорта превышает лимит 50 столбцов")
        headers = [str(h).strip() if h is not None else "" for h in header_row]
        out = []
        for cells in row_iterator:
            if any(cell.data_type == "f" for cell in cells):
                raise ValidationError("Формулы в import-файле запрещены")
            row = [cell.value for cell in cells]
            if any(value is not None for value in row):
                out.append(dict(zip(headers, [("" if value is None else str(value).strip()) for value in row])))
                if len(out) > MAX_IMPORT_ROWS:
                    raise ValidationError("Файл импорта превышает лимит 5 000 строк")
        return headers, out
    text = _decode_csv(raw)
    sample = text[:2048]
    first_line = sample.splitlines()[0] if sample.splitlines() else ""
    if ";" in first_line:
        delimiter = ";"
    elif "\t" in first_line:
        delimiter = "\t"
    else:
        try:
            delimiter = csv.Sniffer().sniff(sample, delimiters=",").delimiter
        except csv.Error:
            delimiter = ","
    reader = csv.DictReader(io.StringIO(text), delimiter=delimiter)
    headers = [h.strip() for h in (reader.fieldnames or [])]
    if len(headers) > MAX_IMPORT_COLUMNS:
        raise ValidationError("Файл импорта превышает лимит 50 столбцов")
    rows = []
    for row in reader:
        rows.append({(key or "").strip(): (value or "").strip() for key, value in row.items()})
        if len(rows) > MAX_IMPORT_ROWS:
            raise ValidationError("Файл импорта превышает лимит 5 000 строк")
    return headers, rows


def _apply_mapping(row: dict, mapping: dict) -> dict:
    """mapping: {source_header: canonical_field}. Splits `name` into first/last."""
    data = {}
    for src, canon in mapping.items():
        data[canon] = row.get(src, "").strip()
    if data.get("name") and not (data.get("first_name") or data.get("last_name")):
        parts = data["name"].split(maxsplit=1)
        data["last_name"] = parts[0]
        data["first_name"] = parts[1] if len(parts) > 1 else ""
    return data


def _dedup_key(data):
    if data.get("email"):
        return ("email", data["email"].lower())
    return ("name_phone", data.get("last_name", "").lower(),
            data.get("first_name", "").lower(),
            data.get("parent_phone") or data.get("phone", ""))


def preview(headers, rows, mapping):
    """Validate + classify each row (new / duplicate / error). No DB writes."""
    rows = prepare_rows("clients", headers, rows, mapping)["rows"]
    seen = set()
    seen_record_ids = set()
    existing_emails = set(e.lower() for e in
                          Student.objects.exclude(email="").values_list("email", flat=True))
    result = []
    for i, row in enumerate(rows, start=2):  # row 1 = header
        data = dict(row)
        if data.get("name") and not (data.get("first_name") or data.get("last_name")):
            parts = str(data["name"]).split(maxsplit=1)
            data["last_name"] = parts[0]
            data["first_name"] = parts[1] if len(parts) > 1 else ""
        pr = PreviewRow(index=i, data=data)
        record_id = str(data.get("record_id") or "").strip()
        if record_id:
            if record_id in seen_record_ids:
                pr.status = ERROR
                pr.errors.append("Повторяющийся Internal ID в файле")
            seen_record_ids.add(record_id)
        if not data.get("last_name") and not data.get("first_name"):
            pr.status = ERROR
            pr.errors.append("Отсутствует имя/фамилия")
        if data.get("email"):
            try:
                validate_email(data["email"])
            except ValidationError:
                pr.status = ERROR
                pr.errors.append("Некорректный email")
        key = _dedup_key(data)
        is_dup = (data.get("email", "").lower() in existing_emails) or (key in seen)
        if pr.status != ERROR and is_dup:
            pr.status = DUPLICATE
            pr.errors.append("Дубликат (email/ФИО+телефон уже есть)")
        group = None
        raw_group_id = data.get("group_id")
        if raw_group_id not in (None, ""):
            try:
                group = Group.objects.filter(pk=int(raw_group_id)).first()
            except (TypeError, ValueError):
                pr.status = ERROR
                pr.errors.append("Некорректный Group internal ID")
        if group is None and data.get("group_name"):
            group = Group.objects.filter(name__iexact=str(data["group_name"]).strip()).first()
        if group:
            pr.resolved["group_id"] = group.id
            pr.resolved["group_reason"] = (
                "Совпал стабильный internal ID" if str(group.id) == str(raw_group_id)
                else "Совпало название группы")
        seen.add(key)
        result.append(pr)
    return result


@transaction.atomic
def commit(preview_rows, *, actor=None, source_name="", create_missing_groups=True,
           create_subscriptions=True, batch=None):
    """Apply NEW rows atomically. ERROR/DUPLICATE rows are skipped.
    Records an ImportBatch so the whole import can be rolled back later."""
    if batch is None:
        batch = ImportBatch(
            created_by=actor,
            source_name=source_name,
            kind=ImportKind.CLIENTS,
            status=ImportBatchStatus.COMMITTED,
            committed_at=timezone.now(),
            rows_total=len(preview_rows),
        )
    else:
        batch.created_by = actor
        batch.source_name = source_name or batch.source_name
    created_students, created_groups, created_parents, created_subscriptions = [], [], [], []
    group_cache = {}

    for pr in preview_rows:
        if pr.status != NEW:
            continue
        d = pr.data
        group = None
        raw_group_id = d.get("group_id")
        if raw_group_id not in (None, ""):
            try:
                group = Group.objects.filter(pk=int(raw_group_id)).first()
            except (TypeError, ValueError):
                group = None
        gname = str(d.get("group_name") or d.get("group") or "").strip()
        if gname:
            if group is not None:
                group_cache[gname] = group
            elif gname in group_cache:
                group = group_cache[gname]
            else:
                group = Group.objects.filter(name=gname).first()
                if group is None:
                    if not create_missing_groups:
                        raise ValidationError(f"Группа не найдена: {gname}")
                    group = Group.objects.create(name=gname)
                    created_groups.append(group.id)
                group_cache[gname] = group

        # Family = ParentAccount keyed by phone (DECISIONS.md #3): siblings sharing
        # one phone join the same family instead of spawning duplicate accounts.
        phone = str(d.get("parent_phone") or d.get("phone") or "").strip()
        parent = None
        raw_parent_id = d.get("parent_record_id")
        if raw_parent_id not in (None, ""):
            try:
                parent = ParentAccount.objects.filter(pk=int(raw_parent_id)).first()
            except (TypeError, ValueError):
                parent = None
        if parent is None and phone:
            parent = ParentAccount.objects.filter(phone=phone).first()
        if parent is None:
            requested_username = str(d.get("parent_username") or "").strip()
            base = (requested_username or d.get("parent_email") or d.get("email") or phone or
                    f'{d.get("last_name","")}{d.get("first_name","")}').strip() or "client"
            username = (requested_username or f"imp_{base}")[:150]
            n = 1
            while User.objects.filter(username=username).exists():
                n += 1
                username = f"{requested_username or f'imp_{base}'}_{n}"[:150]
            puser = User.objects.create_user(
                username=username, role="parent",
                first_name=str(d.get("parent_first_name") or d.get("first_name") or "").strip(),
                last_name=str(d.get("parent_last_name") or d.get("last_name") or "").strip(),
                email=str(d.get("parent_email") or d.get("email") or "").strip())
            # Imported clients keep their domain identity and history, but portal
            # access is enabled only through the one-time activation flow.
            puser.set_unusable_password()
            puser.save(update_fields=["password"])
            parent_kwargs = {
                "user": puser,
                "phone": phone,
                "email": str(d.get("parent_email") or d.get("email") or "").strip(),
                "preferred_language": str(d.get("preferred_language") or "pl").strip(),
            }
            if (raw_parent_id not in (None, "") and
                    not ParentAccount.objects.filter(pk=raw_parent_id).exists()):
                parent_kwargs["id"] = int(raw_parent_id)
            parent = ParentAccount.objects.create(**parent_kwargs)
            created_parents.append(parent.id)

        birth_date = None
        if d.get("birth_date"):
            try:
                birth_date = date.fromisoformat(str(d["birth_date"]).strip())
            except ValueError as exc:
                raise ValidationError(f"Некорректная дата рождения: {d['birth_date']}") from exc
        def import_bool(value, default=False):
            if value in (None, ""):
                return default
            key = str(value).strip().casefold()
            if key in {"true", "1", "yes", "да", "tak"}:
                return True
            if key in {"false", "0", "no", "нет", "nie"}:
                return False
            raise ValidationError(f"Некорректное boolean-значение: {value}")
        student_kwargs = {
            "parent": parent,
            "group": group,
            "first_name": str(d.get("first_name") or "").strip(),
            "last_name": str(d.get("last_name") or "").strip(),
            "birth_date": birth_date,
            "email": str(d.get("email") or "").strip(),
            "is_account_holder": import_bool(d.get("is_account_holder"), False),
            "medical_info": str(d.get("medical_info") or ""),
            "contraindications": str(d.get("contraindications") or ""),
            "emergency_contact_name": str(d.get("emergency_contact_name") or ""),
            "emergency_contact_phone": str(d.get("emergency_contact_phone") or ""),
            "is_active": import_bool(d.get("is_active"), True),
            "admin_comments": str(d.get("admin_comments") or ""),
        }
        raw_student_id = d.get("record_id")
        if (raw_student_id not in (None, "") and
                not Student.objects.filter(pk=raw_student_id).exists()):
            student_kwargs["id"] = int(raw_student_id)
        student = Student.objects.create(**student_kwargs)
        created_students.append(student.id)
        audit(actor, "client.created", student, {"source": "import"})

        stype_name = str(d.get("subscription_type_name") or d.get("subscription") or "").strip()
        if create_subscriptions and stype_name:
            stype = SubscriptionType.objects.filter(name=stype_name).first()
            if stype:
                subscription = create_subscription(student=student, subscription_type=stype,
                                                   start_date=date.today(), created_by=actor)
                created_subscriptions.append(subscription.id)

    batch.created_student_ids = created_students
    batch.created_group_ids = created_groups
    batch.created_parent_ids = created_parents
    batch.created_subscription_ids = created_subscriptions
    batch.rows_imported = len(created_students)
    batch.status = ImportBatchStatus.COMMITTED
    batch.committed_at = timezone.now()
    batch.input_data = {}
    batch.preview_expires_at = None
    batch.result = {
        "rows_total": batch.rows_total,
        "rows_imported": batch.rows_imported,
    }
    batch.save()
    return batch


def rollback_preview(batch: ImportBatch):
    """Return blockers without changing data; a rollback may never delete later work."""
    student_ids = list(batch.created_student_ids or [])
    known_subscription_ids = set(batch.created_subscription_ids or [])
    blockers = []
    for student in Student.objects.filter(id__in=student_ids):
        counts = {
            "attendance": student.attendance.count(),
            "payments": student.payments.count(),
            "charges": student.charges.count(),
            "session_participations": student.session_participations.count(),
            "waitlist_entries": student.waitlist_entries.count(),
            "later_subscriptions": student.subscriptions.exclude(id__in=known_subscription_ids).count(),
        }
        active = {key: value for key, value in counts.items() if value}
        if active:
            blockers.append({"student_id": student.id, "student": student.full_name, "dependencies": active})
    return {
        "batch_id": batch.id,
        "can_rollback": not blockers and not batch.is_rolled_back,
        "blockers": blockers,
        "will_delete": {
            "students": len(student_ids),
            "parents": len(batch.created_parent_ids or []),
            "groups": len(batch.created_group_ids or []),
            "subscriptions": len(batch.created_subscription_ids or []),
        },
    }


@transaction.atomic
def rollback(batch: ImportBatch):
    """Undo a committed import only if no later operational/financial data exists."""
    if batch.is_rolled_back:
        raise ValidationError("Импорт уже откачен")
    preview = rollback_preview(batch)
    if preview["blockers"]:
        raise ValidationError("Откат заблокирован: импортированные клиенты уже получили новые данные")
    # Delete created students explicitly (some may hang off a reused family parent
    # that must survive), then drop the parents this batch created (cascades users).
    with allow_subscription_history_delete():
        Student.objects.filter(id__in=batch.created_student_ids).delete()
    user_ids = list(ParentAccount.objects.filter(id__in=batch.created_parent_ids)
                    .values_list("user_id", flat=True))
    User.objects.filter(id__in=user_ids).delete()
    Group.objects.filter(id__in=batch.created_group_ids, students__isnull=True).delete()
    batch.is_rolled_back = True
    batch.status = ImportBatchStatus.ROLLED_BACK
    batch.save(update_fields=["is_rolled_back", "status"])
    return batch
