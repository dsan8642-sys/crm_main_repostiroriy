"""Seal the owner-approved migration inputs against the live target database."""

import json
from datetime import date, datetime
from pathlib import Path

from django.core.management.base import BaseCommand, CommandError
from openpyxl import load_workbook

from dataio.legacy_migration import file_sha256, production_snapshot
from students.models import Student


APPROVED_FAMILY_CREATES = {
    "663888", "756027", "661221", "660278", "660279",
    "673360", "676825", "731356", "731529",
}


def _clean(value):
    return str(value or "").strip()


def _records(sheet):
    rows = list(sheet.iter_rows(values_only=True))
    headers = [_clean(value) for value in rows[0]]
    return [dict(zip(headers, row)) for row in rows[1:] if any(value is not None for value in row)]


def _iso(value):
    if value in (None, ""):
        return None
    if isinstance(value, datetime):
        return value.date().isoformat()
    if isinstance(value, date):
        return value.isoformat()
    return _clean(value)


def _identity(student):
    return {
        "first_name": student.first_name,
        "last_name": student.last_name,
        "birth_date": _iso(student.birth_date),
    }


def build_family_create_client(legacy_id, mapped, canonical_student):
    fields = mapped.get("create_fields") or {}
    first_name = _clean(fields.get("first_name"))
    last_name = _clean(fields.get("last_name"))
    if not first_name or not last_name:
        raise CommandError(f"Family create {legacy_id} requires an approved Latin first and last name.")
    return {
        "legacy_id": legacy_id,
        "approved": True,
        "action": "create",
        "parent_target_id": canonical_student.parent_id,
        "fields": {
            "first_name": first_name,
            "last_name": last_name,
            "birth_date": _iso(fields.get("birth_date")),
            "email": "",
            "phone": "",
            "parent_email": "",
            "is_account_holder": False,
        },
        "allow_name_overwrite": False,
    }


class Command(BaseCommand):
    help = "Seal approved legacy mappings and balances into an executable manifest."

    def add_arguments(self, parser):
        parser.add_argument("--source-workbook", required=True)
        parser.add_argument("--mapping-json", required=True)
        parser.add_argument("--balance-review-json", required=True)
        parser.add_argument("--run-id", required=True)
        parser.add_argument("--output", required=True)

    def handle(self, *args, **options):
        source_path = Path(options["source_workbook"])
        mapping = json.loads(Path(options["mapping_json"]).read_text(encoding="utf-8"))
        balances_review = json.loads(Path(options["balance_review_json"]).read_text(encoding="utf-8"))
        book = load_workbook(source_path, read_only=True, data_only=True)
        matches = _records(book["Client_Matches"])
        contacts = {str(row["legacy_id"]): row for row in _records(book["Contact_Changes"])}
        mapping_by_id = {str(row["legacy_id"]): row for row in mapping["rows"]}
        approved = [row for row in matches if row.get("review_decision") == "Approved"]
        excluded = [row for row in matches if row.get("review_decision") == "Excluded"]

        if len(approved) != 293:
            raise CommandError(f"Expected 293 approved legacy IDs, got {len(approved)}.")
        if {str(row["legacy_id"]) for row in excluded} != {"603302", "605680"}:
            raise CommandError("The Sosnov exclusions are not exact.")

        clients = []
        for match in approved:
            legacy_id = str(match["legacy_id"])
            mapped = mapping_by_id.get(legacy_id)
            if not mapped or len(mapped.get("candidates", [])) != 1:
                raise CommandError(f"Legacy ID {legacy_id} has no unique stable target.")
            method = _clean(mapped.get("method"))
            alias_of = method.removeprefix("alias_of:")
            if method.startswith("alias_of:"):
                clients.append({"legacy_id": legacy_id, "approved": True, "action": "alias", "alias_of": alias_of})
                continue

            if method.startswith("create_under_parent_of:"):
                canonical_id = method.removeprefix("create_under_parent_of:")
                canonical = mapping_by_id.get(canonical_id)
                if not canonical or len(canonical.get("candidates", [])) != 1:
                    raise CommandError(f"Family create {legacy_id} has no unique canonical target.")
                if _clean(canonical.get("method")).startswith(("alias_of:", "create_under_parent_of:")):
                    raise CommandError(f"Family create {legacy_id} must reference a canonical update target.")
                if mapped["candidates"][0]["student_id"] != canonical["candidates"][0]["student_id"]:
                    raise CommandError(f"Family create {legacy_id} target differs from its canonical family target.")
                canonical_student = Student.objects.filter(
                    id=canonical["candidates"][0]["student_id"]
                ).select_related("parent").first()
                if canonical_student is None:
                    raise CommandError(f"Family create {legacy_id} canonical Student.id is absent.")
                clients.append(build_family_create_client(legacy_id, mapped, canonical_student))
                continue

            student = Student.objects.filter(id=mapped["candidates"][0]["student_id"]).select_related("parent").first()
            if student is None:
                raise CommandError(f"Stable Student.id is absent: {mapped['candidates'][0]['student_id']} for {legacy_id}.")
            contact = contacts.get(legacy_id, {})
            allow_name = contact.get("name_action") == "PROPOSE_LATIN"
            fields = {
                "first_name": _clean(contact.get("proposed_first_name")) or student.first_name,
                "last_name": _clean(contact.get("proposed_last_name")) or student.last_name,
                "birth_date": _iso(student.birth_date),
                "email": _clean(contact.get("legacy_email")) or _clean(match.get("legacy_email")),
                "phone": _clean(contact.get("legacy_phone")) or _clean(match.get("legacy_phone")),
                "parent_email": _clean(contact.get("legacy_email")) or _clean(match.get("legacy_email")),
                "is_account_holder": student.is_account_holder,
            }
            clients.append({
                "legacy_id": legacy_id,
                "approved": True,
                "action": "update",
                "target_student_id": student.id,
                "expected_target": _identity(student),
                "fields": fields,
                "allow_name_overwrite": allow_name,
            })

        family_create_ids = {row["legacy_id"] for row in clients if row["action"] == "create"}
        if family_create_ids != APPROVED_FAMILY_CREATES:
            raise CommandError("Approved family participant creations are not exact.")

        balance_by_id = {str(row["legacy_id"]): row for row in balances_review}
        balances = []
        client_by_id = {str(row["legacy_id"]): row for row in clients}
        for legacy_id, review in balance_by_id.items():
            client = client_by_id.get(legacy_id)
            if client is None or client["action"] == "alias":
                raise CommandError(f"Balance {legacy_id} has no canonical approved client.")
            student = Student.objects.get(id=client["target_student_id"])
            current = int(review["current_sessions"])
            if int(review["student_id"]) != student.id or current != int(review["expected_final_sessions"]) - int(review["legacy_adjustment"]):
                raise CommandError(f"Balance {legacy_id} does not reconcile to its final review.")
            balances.append({
                "legacy_id": legacy_id,
                "approved": True,
                "current_sessions": current,
                "legacy_adjustment_sessions": int(review["legacy_adjustment"]),
                "expected_final_sessions": int(review["expected_final_sessions"]),
                "pln_remainder": 0,
                "evidence": _clean(review.get("evidence")),
            })

        manifest = {
            "schema_version": 1,
            "run_id": options["run_id"],
            "source_workbook_sha256": file_sha256(source_path),
            "target_snapshot_sha256": "",
            "expected_approved_legacy_count": 293,
            "subscription_type": {"name": "8 тренировок", "sessions_count": 8, "duration_days": 31},
            "clients": clients,
            "new_clients": [
                {
                    "legacy_id": "gsheet:duda_aleksander", "approved": True, "action": "create",
                    "fields": {"first_name": "Aleksander", "last_name": "Duda", "birth_date": None,
                               "email": "", "phone": "", "parent_email": "", "is_account_holder": False},
                    "allow_name_overwrite": False,
                    "new_parent": {"first_name": "", "last_name": "", "phone": "", "email": ""},
                },
                {
                    "legacy_id": "gsheet:duda_jakub", "approved": True, "action": "create",
                    "fields": {"first_name": "Jakub", "last_name": "Duda", "birth_date": None,
                               "email": "", "phone": "", "parent_email": "", "is_account_holder": False},
                    "allow_name_overwrite": False,
                    "new_parent": {"first_name": "", "last_name": "", "phone": "", "email": ""},
                },
            ],
            "balances": balances + [
                {"legacy_id": "gsheet:duda_aleksander", "approved": True, "current_sessions": 0,
                 "legacy_adjustment_sessions": 2, "expected_final_sessions": 2, "pln_remainder": 0,
                 "evidence": "Approved payment evidence; separate uninvited family account"},
                {"legacy_id": "gsheet:duda_jakub", "approved": True, "current_sessions": 0,
                 "legacy_adjustment_sessions": 2, "expected_final_sessions": 2, "pln_remainder": 0,
                 "evidence": "Approved payment evidence; separate uninvited family account"},
            ],
            "exclusions": [{"legacy_id": "603302", "approved": True}, {"legacy_id": "605680", "approved": True}],
        }
        manifest["target_snapshot_sha256"] = production_snapshot(manifest)
        output = Path(options["output"])
        output.parent.mkdir(parents=True, exist_ok=True)
        output.write_text(json.dumps(manifest, ensure_ascii=False, indent=2, sort_keys=True) + "\n", encoding="utf-8")
        self.stdout.write(self.style.SUCCESS(f"Sealed manifest written: {output}"))
