import io
import json
from datetime import date

from django.core.exceptions import ValidationError
from django.core.files.uploadedfile import SimpleUploadedFile
from django.db import transaction
from django.test import Client, TestCase
from django.urls import Resolver404, resolve
from openpyxl import Workbook, load_workbook

from attendance.models import AttendanceRecord, AttendanceStatus
from billing.models import Payment, PaymentMethod, PaymentSource, PaymentStatus
from catalog.models import Group
from dataio import exports
from dataio.contracts import CONTRACTS
from dataio.importer import MAX_IMPORT_BYTES, MAX_IMPORT_ROWS, parse_source
from dataio.matching import match_student, normalize_email, normalize_phone
from dataio.models import ImportBatch, ImportKind
from scheduling.models import Session
from scheduling.services import create_session
from students.models import Student

from . import factories as f


class ExportImportContractRegressionTest(TestCase):
    def setUp(self):
        self.admin = f.make_admin(username="roundtrip_admin")
        self.client = Client()
        self.client.force_login(self.admin)

    def test_own_clients_csv_is_recognized_without_manual_mapping(self):
        student = f.make_student(
            first="Ирина",
            last="Раундтрип",
            email="roundtrip.client@example.test",
        )
        filename, content = exports.export_entity("clients", "csv")

        response = self.client.post(
            "/api/admin/import/clients/preview/",
            {"file": SimpleUploadedFile(filename, content, content_type="text/csv")},
        )

        self.assertEqual(response.status_code, 200)
        rows = response.json()["rows"]
        self.assertEqual(len(rows), 1)
        own_row = rows[0]
        self.assertEqual(own_row["data"].get("email"), student.email)
        self.assertEqual(own_row["status"], "duplicate")
        self.assertEqual(own_row["errors"], ["Дубликат (email/ФИО+телефон уже есть)"])

    def test_every_standard_export_has_an_import_preview_route(self):
        missing = []
        for entity in exports.DATASETS:
            try:
                resolve(f"/api/admin/import/{entity}/preview/")
            except Resolver404:
                missing.append(entity)
        self.assertEqual(missing, [])

    def test_export_contract_import_kind_and_routes_stay_in_parity(self):
        expected = set(exports.DATASETS)
        self.assertEqual(set(CONTRACTS), expected)
        self.assertEqual(set(ImportKind.values), expected)
        for entity in expected:
            self.assertIsNotNone(resolve(f"/api/admin/import/{entity}/preview/"))
            self.assertIsNotNone(resolve(f"/api/admin/import/{entity}/commit/"))


class StandardDatasetRoundTripTest(TestCase):
    class _RollbackSource(Exception):
        pass

    def setUp(self):
        self.admin = f.make_admin(username="roundtrip_restore_admin")
        self.client = Client()
        self.client.force_login(self.admin)

    def _source_exports(self, fmt):
        captured = {}
        with self.assertRaises(self._RollbackSource):
            with transaction.atomic():
                trainer = f.make_trainer(username="roundtrip_coach")
                trainer.user.first_name = "Тест"
                trainer.user.last_name = "Тренер"
                trainer.user.email = "roundtrip.coach@example.test"
                trainer.user.save(update_fields=["first_name", "last_name", "email"])
                trainer.phone = "+48555101010"
                trainer.save(update_fields=["phone"])
                group = Group.objects.create(
                    name="Round-trip группа",
                    description="Описание",
                    default_trainer=trainer,
                    price_minor=7650,
                    currency="PLN",
                    default_capacity=9,
                    color_key="ocean",
                )
                parent = f.make_parent(username="roundtrip_family", phone="+48555202020")
                parent.user.first_name = "Анна"
                parent.user.last_name = "Родитель"
                parent.user.save(update_fields=["first_name", "last_name"])
                parent.email = "roundtrip.family@example.test"
                parent.save(update_fields=["email"])
                student = Student.objects.create(
                    parent=parent,
                    group=group,
                    first_name="Ирина",
                    last_name="Пловец",
                    birth_date=date(2014, 5, 6),
                    email="roundtrip.student@example.test",
                    medical_info="Синтетические данные",
                    emergency_contact_name="Тест Контакт",
                    emergency_contact_phone="+48555303030",
                    admin_comments="Round-trip comment",
                )
                payment = Payment.objects.create(
                    student=student,
                    amount_minor=12345,
                    currency="PLN",
                    paid_at=date(2026, 4, 3),
                    method=PaymentMethod.TRANSFER,
                    reference_id="roundtrip-ref-001",
                    comment="Round-trip payment",
                    status=PaymentStatus.CONFIRMED,
                    source=PaymentSource.ADMIN,
                    created_by=self.admin,
                    confirmed_by=self.admin,
                )
                session = create_session(
                    trainer=trainer,
                    start_at=f.dt(2026, 4, 4, 9),
                    duration_minutes=60,
                    location="Round-trip pool",
                    max_participants=9,
                    group=group,
                    price_minor=7650,
                    currency="PLN",
                )
                AttendanceRecord.objects.create(
                    session=session,
                    student=student,
                    status=AttendanceStatus.PRESENT,
                    comment="Round-trip attendance",
                    marked_by=self.admin,
                    financial_effects_enabled=False,
                )
                self.assertEqual(payment.student, student)
                for entity in ("trainers", "groups", "clients", "payments", "attendance"):
                    captured[entity] = exports.export_entity(entity, fmt)
                raise self._RollbackSource
        return captured

    def _restore(self, fmt):
        captured = self._source_exports(fmt)
        for entity in ("trainers", "groups", "clients", "payments", "attendance"):
            filename, content = captured[entity]
            preview = self.client.post(
                f"/api/admin/import/{entity}/preview/",
                {"file": SimpleUploadedFile(filename, content)},
            )
            self.assertEqual(preview.status_code, 200, preview.content)
            payload = preview.json()
            selected = [
                row["index"] for row in payload["rows"]
                if row["status"] not in {"error", "duplicate"}
            ]
            self.assertTrue(selected, payload)
            commit = self.client.post(
                f"/api/admin/import/{entity}/commit/",
                data=json.dumps({
                    "batch_id": payload["batch_id"],
                    "selected_indices": selected,
                }),
                content_type="application/json",
            )
            self.assertEqual(commit.status_code, 201, commit.content)

        group = Group.objects.get(name="Round-trip группа")
        student = Student.objects.get(email="roundtrip.student@example.test")
        payment = Payment.objects.get(reference_id="roundtrip-ref-001")
        attendance = AttendanceRecord.objects.get(student=student)
        session = attendance.session
        self.assertEqual(group.default_trainer.user.username, "roundtrip_coach")
        self.assertEqual(group.price_minor, 7650)
        self.assertEqual(student.group, group)
        self.assertEqual(student.birth_date, date(2014, 5, 6))
        self.assertEqual(student.medical_info, "Синтетические данные")
        self.assertEqual(payment.student, student)
        self.assertEqual(payment.amount_minor, 12345)
        self.assertEqual(payment.method, PaymentMethod.TRANSFER)
        self.assertEqual(payment.status, PaymentStatus.CONFIRMED)
        self.assertEqual(payment.comment, "Round-trip payment")
        self.assertEqual(session.group, group)
        self.assertEqual(session.trainer, group.default_trainer)
        self.assertEqual(session.location, "Round-trip pool")
        self.assertEqual(attendance.status, AttendanceStatus.PRESENT)
        self.assertEqual(attendance.comment, "Round-trip attendance")
        self.assertFalse(attendance.financial_effects_enabled)

    def test_csv_round_trip_all_standard_datasets(self):
        self._restore("csv")

    def test_xlsx_round_trip_all_standard_datasets(self):
        self._restore("xlsx")


class StagingWorkflowIntegrationTest(TestCase):
    def setUp(self):
        self.admin = f.make_admin(username="staging_admin")
        self.client = Client()
        self.client.force_login(self.admin)
        self.auto_client = f.make_student(
            first="Авто", last="Клиент", email="auto.match@example.test")
        self.manual_client = f.make_student(
            first="Ручной", last="Клиент", email="manual.match@example.test")

    def _payment_file(self, rows=None):
        rows = rows or [
            "auto.match@example.test;75.25;PLN;2026-03-20;cash;confirmed;staged;stage-ref-1"
        ]
        content = (
            "Клиент;Сумма;Валюта;Дата;Способ;Статус;Комментарий;Reference ID\r\n"
            + "\r\n".join(rows) + "\r\n"
        ).encode("utf-8")
        return SimpleUploadedFile("payments.csv", content, content_type="text/csv"), content

    def test_manual_client_override_is_server_owned_reported_and_duplicate_file_warns(self):
        upload, content = self._payment_file()
        preview = self.client.post(
            "/api/admin/import/payments/preview/", {"file": upload})
        self.assertEqual(preview.status_code, 200, preview.content)
        payload = preview.json()
        batch_id = payload["batch_id"]
        self.assertFalse(Payment.objects.exists(), "preview must not write domain records")
        self.assertEqual(payload["rows"][0]["resolved"]["student_id"], self.auto_client.id)
        self.assertEqual(payload["source_samples"]["Клиент"], "auto.match@example.test")

        protected = self.client.patch(
            f"/api/admin/import/payments/{batch_id}/rows/2/",
            data=json.dumps({"data": {"source": "api"}}),
            content_type="application/json")
        self.assertEqual(protected.status_code, 400)

        edited = self.client.patch(
            f"/api/admin/import/payments/{batch_id}/rows/2/",
            data=json.dumps({"data": {"amount": "76.25"}}),
            content_type="application/json")
        self.assertEqual(edited.status_code, 200, edited.content)
        self.assertEqual(edited.json()["row"]["data"]["amount"], "76.25")

        patched = self.client.patch(
            f"/api/admin/import/payments/{batch_id}/rows/2/",
            data=json.dumps({"relations": {"client_id": self.manual_client.id}}),
            content_type="application/json")
        self.assertEqual(patched.status_code, 200, patched.content)
        self.assertEqual(patched.json()["row"]["resolved"]["student_id"], self.manual_client.id)
        self.assertEqual(patched.json()["row"]["data"]["amount"], "76.25")
        self.assertEqual(
            patched.json()["row"]["resolved"]["matching_reason"], "Клиент выбран вручную")

        committed = self.client.post(
            "/api/admin/import/payments/commit/",
            data=json.dumps({"batch_id": batch_id, "selected_indices": [2]}),
            content_type="application/json")
        self.assertEqual(committed.status_code, 201, committed.content)
        payment = Payment.objects.get(reference_id="stage-ref-1")
        self.assertEqual(payment.student, self.manual_client)
        self.assertEqual(payment.amount_minor, 7625)
        batch = ImportBatch.objects.get(pk=batch_id)
        self.assertEqual(
            batch.result["manual_corrections"]["2"]["data"]["amount"], "76.25")
        self.assertEqual(
            batch.result["manual_corrections"]["2"]["relations"]["client_id"],
            str(self.manual_client.id))
        self.assertEqual(batch.result["rollback_strategy"]["kind"], "compensating_only")
        self.assertEqual(batch.result["report_rows"][0]["created_id"], payment.id)

        report = self.client.get(f"/api/admin/import/batches/{batch_id}/report/csv/")
        self.assertEqual(report.status_code, 200)
        self.assertTrue(report.content.startswith(b"\xef\xbb\xbf"))
        self.assertIn(b"stage-ref-1", report.content)

        other = Client()
        other.force_login(f.make_admin(username="other_report_admin"))
        self.assertEqual(
            other.get(f"/api/admin/import/batches/{batch_id}/report/csv/").status_code,
            400)

        repeated = self.client.post(
            "/api/admin/import/payments/preview/",
            {"file": SimpleUploadedFile("payments.csv", content, content_type="text/csv")})
        self.assertEqual(repeated.status_code, 200)
        self.assertTrue(repeated.json()["duplicate_file"])

    def test_bulk_assignment_resolves_unmatched_rows(self):
        upload, _content = self._payment_file([
            "missing-one@example.test;10;PLN;2026-03-21;cash;confirmed;;bulk-ref-1",
            "missing-two@example.test;20;PLN;2026-03-22;card;confirmed;;bulk-ref-2",
        ])
        preview = self.client.post(
            "/api/admin/import/payments/preview/", {"file": upload}).json()
        self.assertEqual([row["status"] for row in preview["rows"]], ["error", "error"])
        bulk = self.client.post(
            f"/api/admin/import/payments/{preview['batch_id']}/rows/bulk/",
            data=json.dumps({
                "indices": [2, 3],
                "relations": {"client_id": self.manual_client.id},
            }), content_type="application/json")
        self.assertEqual(bulk.status_code, 200, bulk.content)
        self.assertEqual([row["status"] for row in bulk.json()["rows"]], ["new", "new"])
        self.assertTrue(all(
            row["resolved"]["matching_confidence"] == "manual"
            for row in bulk.json()["rows"]))

    def test_explicit_create_client_requires_staging_action(self):
        content = (
            "Email клиента;Имя клиента;Фамилия клиента;Телефон клиента;Сумма;Дата;Способ;Reference ID\r\n"
            "new.person@example.test;Новый;Клиент;+48555999888;30;2026-03-25;cash;create-ref-1\r\n"
        ).encode("utf-8")
        preview = self.client.post(
            "/api/admin/import/payments/preview/",
            {"file": SimpleUploadedFile("new-client.csv", content)}).json()
        self.assertEqual(preview["rows"][0]["status"], "error")
        self.assertFalse(Student.objects.filter(email="new.person@example.test").exists())
        patched = self.client.patch(
            f"/api/admin/import/payments/{preview['batch_id']}/rows/2/",
            data=json.dumps({"relations": {"create_client": "true"}}),
            content_type="application/json")
        self.assertEqual(patched.status_code, 200, patched.content)
        self.assertEqual(patched.json()["row"]["status"], "new")
        committed = self.client.post(
            "/api/admin/import/payments/commit/",
            data=json.dumps({"batch_id": preview["batch_id"], "selected_indices": [2]}),
            content_type="application/json")
        self.assertEqual(committed.status_code, 201, committed.content)
        student = Student.objects.get(email="new.person@example.test")
        self.assertTrue(Payment.objects.filter(student=student, reference_id="create-ref-1").exists())

    def test_group_create_update_skip_modes_and_safe_rollback(self):
        group = Group.objects.create(
            name="Режимы импорта", description="До", default_capacity=8)
        content = (
            "Группа;Описание;Вместимость;Валюта;Активна\r\n"
            "Режимы импорта;После;12;PLN;true\r\n"
        ).encode("utf-8")

        create_only = self.client.post(
            "/api/admin/import/groups/preview/", {
                "file": SimpleUploadedFile("groups.csv", content),
                "import_mode": "create_only",
            }).json()
        self.assertEqual(create_only["rows"][0]["status"], "duplicate")
        group.refresh_from_db()
        self.assertEqual(group.description, "До")

        update = self.client.post(
            "/api/admin/import/groups/preview/", {
                "file": SimpleUploadedFile("groups.csv", content),
                "import_mode": "update_existing",
            }).json()
        self.assertEqual(update["rows"][0]["status"], "update")
        self.assertEqual(
            update["rows"][0]["resolved"]["changes"]["description"],
            {"old": "До", "new": "После"})
        committed = self.client.post(
            "/api/admin/import/groups/commit/",
            data=json.dumps({"batch_id": update["batch_id"], "selected_indices": [2]}),
            content_type="application/json")
        self.assertEqual(committed.status_code, 201, committed.content)
        self.assertEqual(committed.json()["updated"], 1)
        group.refresh_from_db()
        self.assertEqual((group.description, group.default_capacity), ("После", 12))

        rollback_preview = self.client.get(
            f"/api/admin/import/groups/{update['batch_id']}/rollback/")
        self.assertTrue(rollback_preview.json()["can_rollback"])
        rolled_back = self.client.post(
            f"/api/admin/import/groups/{update['batch_id']}/rollback/",
            data=json.dumps({
                "confirm_batch_id": update["batch_id"],
                "confirm_rollback": True,
            }), content_type="application/json")
        self.assertEqual(rolled_back.status_code, 200, rolled_back.content)
        group.refresh_from_db()
        self.assertEqual((group.description, group.default_capacity), ("До", 8))

        new_content = (
            "Группа;Описание;Вместимость;Валюта;Активна\r\n"
            "Новая upsert группа;Создана;9;PLN;true\r\n"
        ).encode("utf-8")
        upsert = self.client.post(
            "/api/admin/import/groups/preview/", {
                "file": SimpleUploadedFile("groups-upsert.csv", new_content),
                "import_mode": "upsert",
            }).json()
        self.assertEqual(upsert["rows"][0]["status"], "new")


class ImportFileSafetyTest(TestCase):
    def test_xlsx_formula_is_rejected_instead_of_evaluated(self):
        workbook = Workbook()
        sheet = workbook.active
        sheet.append(["Клиент", "Сумма"])
        sheet.append(["client@example.test", "=1+1"])
        buffer = io.BytesIO()
        workbook.save(buffer)
        with self.assertRaisesMessage(ValidationError, "Формулы в import-файле запрещены"):
            parse_source(buffer.getvalue(), "payments.xlsx")

    def test_exported_formula_like_text_is_never_an_xlsx_formula(self):
        student = f.make_student(
            first="Формула", last="Тест", email="formula@example.test")
        student.admin_comments = "=HYPERLINK(\"https://invalid.test\")"
        student.save(update_fields=["admin_comments"])
        _name, content = exports.export_entity("clients", "xlsx")
        workbook = load_workbook(io.BytesIO(content), data_only=False)
        cells = [cell for row in workbook.active.iter_rows() for cell in row]
        matching = [cell for cell in cells if "HYPERLINK" in str(cell.value)]
        self.assertEqual(len(matching), 1)
        self.assertNotEqual(matching[0].data_type, "f")
        self.assertTrue(str(matching[0].value).startswith("'="))

    def test_normalizers_and_ambiguous_family_phone_never_auto_match(self):
        parent = f.make_parent(username="ambiguous_family", phone="+48 555-123-123")
        f.make_student(parent=parent, first="Первый", last="Ребёнок")
        f.make_student(parent=parent, first="Второй", last="Ребёнок")
        self.assertEqual(normalize_email(" User@Example.TEST "), "user@example.test")
        self.assertEqual(normalize_phone("+48 (555) 123-123"), "+48555123123")
        matched = match_student({"client_phone": "+48555123123"})
        self.assertIsNone(matched.student)
        self.assertTrue(matched.ambiguous)
        self.assertEqual(len(matched.candidates), 2)

    def test_corrupt_xlsx_executable_extension_and_size_limit_are_rejected(self):
        with self.assertRaisesMessage(
                ValidationError, "Повреждённый или неподдерживаемый XLSX-файл"):
            parse_source(b"not-a-workbook", "payments.xlsx")
        with self.assertRaisesMessage(
                ValidationError, "Поддерживаются только CSV, XLSX и XLSM"):
            parse_source(b"MZ", "payload.exe")
        with self.assertRaisesMessage(ValidationError, "превышает лимит 5 МБ"):
            parse_source(b"x" * (MAX_IMPORT_BYTES + 1), "clients.csv")


class ImportSchemaValidationTest(TestCase):
    def setUp(self):
        self.admin = f.make_admin(username="schema_admin")
        self.client = Client()
        self.client.force_login(self.admin)

    def test_empty_file_and_unknown_own_schema_are_rejected(self):
        empty = self.client.post(
            "/api/admin/import/clients/preview/",
            {"file": SimpleUploadedFile("empty.csv", b"")})
        self.assertEqual(empty.status_code, 400)

        f.make_student(first="Версия", last="Схемы", email="schema@example.test")
        filename, content = exports.export_entity("clients", "csv")
        text = content.decode("utf-8-sig")
        lines = text.splitlines()
        values = lines[1].split(";")
        values[0] = "999"
        unsupported = (lines[0] + "\r\n" + ";".join(values) + "\r\n").encode("utf-8")
        response = self.client.post(
            "/api/admin/import/clients/preview/",
            {"file": SimpleUploadedFile(filename, unsupported)})
        self.assertEqual(response.status_code, 400)
        self.assertIn(
            "schema_version",
            " ".join(item["message"] for item in response.json()["errors"]["file"]),
        )

    def test_preview_rejects_more_than_row_limit_without_creating_batch(self):
        content = (
            "Фамилия;Имя;Email\r\n"
            + "".join(
                f"Лимит;Строка{i};row{i}@example.test\r\n"
                for i in range(MAX_IMPORT_ROWS + 1)
            )
        ).encode("utf-8")

        response = self.client.post(
            "/api/admin/import/clients/preview/",
            {"file": SimpleUploadedFile("too-many.csv", content)},
        )

        self.assertEqual(response.status_code, 400)
        self.assertIn(
            "5 000 строк",
            " ".join(item["message"] for item in response.json()["errors"]["file"]),
        )
        self.assertFalse(ImportBatch.objects.exists())
