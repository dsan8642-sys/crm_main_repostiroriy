from datetime import date, datetime
from io import BytesIO
from zoneinfo import ZoneInfo

from django.test import Client, TestCase
from openpyxl import load_workbook

from billing.models import Payment, PaymentMethod, PaymentStatus
from scheduling.models import SessionType
from scheduling.services import create_session

from . import factories as f


WARSAW = ZoneInfo("Europe/Warsaw")


class AdminReportsWorld:
    def setUp(self):
        self.admin = f.make_admin(username="reports_admin")
        self.client.force_login(self.admin)
        self.group = f.make_group("Reports group")
        self.student = f.make_student(
            group=self.group,
            first="Report",
            last="Participant",
        )

    def make_trainer(self, username, first, last, *, active=True):
        trainer = f.make_trainer(username=username)
        trainer.user.first_name = first
        trainer.user.last_name = last
        trainer.user.save(update_fields=["first_name", "last_name"])
        trainer.is_active = active
        trainer.save(update_fields=["is_active"])
        return trainer

    def make_session(self, trainer, local_start, session_type, *, cancelled=False):
        session = create_session(
            trainer=trainer,
            start_at=local_start,
            duration_minutes=45,
            location="Reports pool",
            max_participants=8 if session_type == SessionType.GROUP else 2,
            group=self.group if session_type == SessionType.GROUP else None,
            session_type=session_type,
            individual_student=(
                None if session_type == SessionType.GROUP else self.student
            ),
        )
        if cancelled:
            session.is_cancelled = True
            session.save(update_fields=["is_cancelled"])
        return session

class AdminReportsHttpRule(AdminReportsWorld, TestCase):
    def test_session_counts_use_warsaw_dates_effective_trainer_and_active_zero_rows(self):
        scheduled = self.make_trainer("scheduled", "Borys", "Scheduled")
        actual = self.make_trainer("actual", "Anna", "Actual")
        zero = self.make_trainer("zero", "Celia", "Same")
        access_revoked = self.make_trainer("revoked", "Celia", "Same")
        access_revoked.user.is_active = False
        access_revoked.user.save(update_fields=["is_active"])
        inactive = self.make_trainer("inactive", "Daria", "Historical", active=False)

        substituted = self.make_session(
            scheduled,
            datetime(2026, 8, 1, 0, 15, tzinfo=WARSAW),
            SessionType.GROUP,
        )
        substituted.substitute_trainer = actual
        substituted.save(update_fields=["substitute_trainer"])
        self.make_session(
            actual,
            datetime(2026, 8, 2, 10, 0, tzinfo=WARSAW),
            SessionType.INDIVIDUAL,
        )
        self.make_session(
            actual,
            datetime(2026, 8, 3, 10, 0, tzinfo=WARSAW),
            SessionType.SPLIT,
        )
        self.make_session(
            actual,
            datetime(2026, 8, 4, 10, 0, tzinfo=WARSAW),
            SessionType.GROUP,
            cancelled=True,
        )
        self.make_session(
            inactive,
            datetime(2026, 8, 5, 10, 0, tzinfo=WARSAW),
            SessionType.GROUP,
        )
        self.make_session(
            actual,
            datetime(2026, 7, 31, 22, 30, tzinfo=WARSAW),
            SessionType.GROUP,
        )

        response = self.client.get(
            "/api/admin/reports/session-counts/",
            {"date_from": "2026-08-01", "date_to": "2026-08-31"},
        )

        self.assertEqual(response.status_code, 200)
        payload = response.json()
        rows = {row["trainer_id"]: row for row in payload["rows"]}
        self.assertEqual(rows[actual.id]["group"], 1)
        self.assertEqual(rows[actual.id]["individual"], 1)
        self.assertEqual(rows[actual.id]["split"], 1)
        self.assertEqual(rows[actual.id]["total"], 3)
        self.assertEqual(rows[scheduled.id]["total"], 0)
        self.assertEqual(rows[zero.id]["total"], 0)
        self.assertEqual(rows[access_revoked.id]["total"], 0)
        self.assertTrue(rows[access_revoked.id]["is_active"])
        self.assertEqual(rows[inactive.id]["group"], 1)
        self.assertFalse(rows[inactive.id]["is_active"])
        self.assertEqual(
            [row["trainer_id"] for row in payload["rows"]],
            [actual.id, scheduled.id, zero.id, access_revoked.id, inactive.id],
        )
        self.assertEqual(
            payload["totals"],
            {"group": 2, "individual": 1, "split": 1, "total": 4},
        )
        self.assertEqual(payload["date_from"], "2026-08-01")
        self.assertEqual(payload["date_to"], "2026-08-31")

        selected = self.client.get(
            "/api/admin/reports/session-counts/",
            {
                "date_from": "2026-08-01",
                "date_to": "2026-08-31",
                "trainer_id": actual.id,
            },
        )
        self.assertEqual(selected.status_code, 200)
        self.assertEqual([row["trainer_id"] for row in selected.json()["rows"]], [actual.id])
        self.assertEqual(selected.json()["totals"]["total"], 3)

    def test_income_report_preserves_legacy_totals_and_adds_cash_non_cash_detail(self):
        cash = Payment.objects.create(
            student=self.student,
            amount_minor=1000,
            currency="PLN",
            paid_at=date(2026, 8, 1),
            method=PaymentMethod.CASH,
            status=PaymentStatus.CONFIRMED,
        )
        card = Payment.objects.create(
            student=self.student,
            amount_minor=2500,
            currency="PLN",
            paid_at=date(2026, 8, 31),
            method=PaymentMethod.CARD,
            status=PaymentStatus.CONFIRMED,
        )
        Payment.objects.create(
            student=self.student,
            amount_minor=9900,
            currency="PLN",
            paid_at=date(2026, 8, 15),
            method=PaymentMethod.CASH,
            status=PaymentStatus.PENDING,
        )
        Payment.objects.create(
            student=self.student,
            amount_minor=700,
            currency="EUR",
            paid_at=date(2026, 8, 15),
            method=PaymentMethod.CASH,
            status=PaymentStatus.CONFIRMED,
        )

        response = self.client.get(
            "/api/admin/reports/income/",
            {
                "date_from": "2026-08-01",
                "date_to": "2026-08-31",
                "currency": "PLN",
                "page": 1,
                "page_size": 1,
            },
        )

        self.assertEqual(response.status_code, 200)
        payload = response.json()
        self.assertEqual(payload["total_minor"], 3500)
        self.assertEqual(payload["cash_minor"], 1000)
        self.assertEqual(payload["non_cash_minor"], 2500)
        self.assertIn("by_group", payload)
        self.assertIn("by_trainer", payload)
        self.assertEqual(payload["pagination"]["total"], 2)
        self.assertEqual(payload["pagination"]["pages"], 2)
        self.assertEqual([row["id"] for row in payload["payments"]], [card.id])
        self.assertEqual(payload["payments"][0]["participant"], self.student.full_name)
        self.assertEqual(payload["payments"][0]["method_label"], "Карта")
        self.assertEqual(payload["available_currencies"], ["PLN", "EUR", "USD"])
        self.assertNotEqual(cash.id, card.id)

    def test_report_validation_and_permissions_are_public_http_behavior(self):
        anonymous = Client()
        for path in (
            "/api/admin/reports/session-counts/",
            "/api/admin/reports/session-counts/xlsx/",
            "/api/admin/reports/income/",
            "/api/admin/reports/income/xlsx/",
        ):
            with self.subTest(path=path):
                response = anonymous.get(
                    path,
                    {"date_from": "2026-08-01", "date_to": "2026-08-31"},
                )
                self.assertEqual(response.status_code, 403)

        invalid_range = self.client.get(
            "/api/admin/reports/session-counts/",
            {"date_from": "2026-08-31", "date_to": "2026-08-01"},
        )
        invalid_trainer = self.client.get(
            "/api/admin/reports/session-counts/",
            {"date_from": "2026-08-01", "date_to": "2026-08-31", "trainer_id": 999999},
        )
        invalid_currency = self.client.get(
            "/api/admin/reports/income/",
            {"date_from": "2026-08-01", "date_to": "2026-08-31", "currency": "BTC"},
        )
        self.assertEqual(invalid_range.status_code, 400)
        self.assertEqual(invalid_trainer.status_code, 400)
        self.assertEqual(invalid_currency.status_code, 400)

        schema = self.client.get("/api/openapi.json").json()
        self.assertIn("/api/admin/reports/session-counts/", schema["paths"])
        self.assertIn("/api/admin/reports/session-counts/xlsx/", schema["paths"])
        self.assertIn("/api/admin/reports/income/xlsx/", schema["paths"])


class AdminReportsXlsxRule(AdminReportsWorld, TestCase):
    def test_report_exports_include_complete_filtered_data(self):
        trainer = self.make_trainer("xlsx_trainer", "Xlsx", "Trainer")
        self.make_session(
            trainer,
            datetime(2026, 8, 12, 10, 0, tzinfo=WARSAW),
            SessionType.GROUP,
        )
        formula_student = f.make_student(
            group=self.group,
            first="Formula",
            last="=DANGEROUS",
        )
        for amount, method in ((1200, PaymentMethod.CASH), (2300, PaymentMethod.TRANSFER)):
            Payment.objects.create(
                student=formula_student,
                amount_minor=amount,
                currency="PLN",
                paid_at=date(2026, 8, 12),
                method=method,
                status=PaymentStatus.CONFIRMED,
            )

        params = {"date_from": "2026-08-01", "date_to": "2026-08-31"}
        sessions_response = self.client.get(
            "/api/admin/reports/session-counts/xlsx/", params,
        )
        income_response = self.client.get(
            "/api/admin/reports/income/xlsx/",
            {**params, "currency": "PLN", "page_size": 1},
        )

        self.assertEqual(sessions_response.status_code, 200)
        self.assertEqual(income_response.status_code, 200)
        self.assertEqual(
            sessions_response["Content-Type"],
            "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )
        sessions_book = load_workbook(BytesIO(sessions_response.content), read_only=True)
        self.assertEqual(sessions_book.sheetnames, ["Итоги", "По тренерам"])
        self.assertEqual(sessions_book["Итоги"]["F2"].value, 1)

        income_book = load_workbook(BytesIO(income_response.content), read_only=True)
        self.assertEqual(income_book.sheetnames, ["Сводка", "Платежи"])
        payment_rows = list(income_book["Платежи"].iter_rows(values_only=True))
        self.assertEqual(len(payment_rows), 3)
        participant_column = payment_rows[0].index("Клиент")
        self.assertTrue(str(payment_rows[1][participant_column]).startswith("'="))
